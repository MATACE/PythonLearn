#!/usr/bin/python
# coding=utf-8

import os
from os import access
import openpyxl
import logging

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import openpyxl.chart

class ExcelOp():
    def __init__(self):
        pass
    def open(self):
        """ Open the excel file
        """        
        os_str = os.getcwd()
        wb = openpyxl.load_workbook(os_str + "/docs//1.xlsx")
        sheets = wb.sheetnames
        for sheet_name in sheets:
            sheet = wb.get_sheet_by_name(sheet_name)
            assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
            # logging.info(type(sheet))
            logging.info("Sheet Name: " + sheet.title)
            for cell_obj in sheet['A1':'C3']:
                for cell in cell_obj:
                    assert isinstance(cell, openpyxl.cell.cell.Cell)
                    logging.info(cell.coordinate)
                    logging.info(cell.value)
    
    def create_grap(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        assert isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet)
        for i in range(1, 11):
            sheet['A' + str(i)] = i

        refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
        seriesObj = openpyxl.chart.Series(refObj, title='First series')
        chartObj = openpyxl.chart.BarChart()
        chartObj.title = 'My Chart'
        chartObj.append(seriesObj)
        sheet.add_chart(chartObj, 'C5')
        wb.save('sampleChart.xlsx')
        